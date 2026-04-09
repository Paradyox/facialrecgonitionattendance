
import os
import json
import pickle
import threading
import tkinter as tk
from datetime import datetime
from pathlib import Path
from tkinter import messagebox, scrolledtext, ttk

import cv2
import face_recognition
import numpy as np
import pandas as pd
from PIL import Image, ImageTk
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill


class FacialRecognitionSystem:

    def __init__(self, data_dir="face_data", attendance_dir="attendance"):
        self.base_dir = Path(__file__).resolve().parent
        self.data_dir = (self.base_dir / data_dir).resolve()
        self.attendance_dir = (self.base_dir / attendance_dir).resolve()
        self.models_dir = (self.base_dir / "models").resolve()
        self.cache_file = self.models_dir / "encodings_cache.pkl"

        self.data_dir.mkdir(exist_ok=True)
        self.attendance_dir.mkdir(exist_ok=True)
        self.models_dir.mkdir(exist_ok=True)

        self.known_face_encodings = []
        self.known_face_names = []

        # One Excel file per app session.
        self.session_started_at = datetime.now()
        self.session_id = self.session_started_at.strftime("%Y%m%d_%H%M%S")
        self.session_attendance_file = self.attendance_dir / f"attendance_session_{self.session_id}.xlsx"
        self.live_json_file = self.base_dir / "attendance_live.json"

        if not self._load_cache():
            self.load_known_faces()

        # Ensure website has a valid starting data snapshot for this session.
        self.export_attendance_json(set())

    def _load_cache(self):
        try:
            if not self.cache_file.exists():
                return False
            with open(self.cache_file, "rb") as f:
                data = pickle.load(f)
            self.known_face_encodings = data.get("encodings", [])
            self.known_face_names = data.get("names", [])
            self._dedupe_known_faces()
            return True
        except Exception:
            return False

    def _dedupe_known_faces(self):
  
        if not self.known_face_names or not self.known_face_encodings:
            return

        grouped = {}
        for name, enc in zip(self.known_face_names, self.known_face_encodings):
            clean = str(name).strip()
            if not clean:
                continue
            grouped.setdefault(clean, []).append(np.asarray(enc, dtype=np.float32))

        dedup_names = []
        dedup_encodings = []
        for name in sorted(grouped.keys()):
            encs = grouped[name]
            if len(encs) == 1:
                dedup_encodings.append(encs[0])
            else:
                dedup_encodings.append(np.mean(encs, axis=0))
            dedup_names.append(name)

        self.known_face_names = dedup_names
        self.known_face_encodings = dedup_encodings

    def _save_cache(self):
        try:
            with open(self.cache_file, "wb") as f:
                pickle.dump(
                    {
                        "encodings": self.known_face_encodings,
                        "names": self.known_face_names,
                    },
                    f,
                )
        except Exception:
            pass

    def load_known_faces(self):
        self.known_face_encodings = []
        self.known_face_names = []

        if not self.data_dir.exists():
            return

        exts = {".jpg", ".jpeg", ".png", ".bmp", ".webp"}

        for person_dir in sorted(self.data_dir.iterdir()):
            if not person_dir.is_dir():
                continue

            person_name = person_dir.name
            person_encodings = []

            for image_file in sorted(person_dir.iterdir()):
                if image_file.suffix.lower() not in exts:
                    continue
                try:
                    image = face_recognition.load_image_file(str(image_file))
                    locations = face_recognition.face_locations(image, model="hog")
                    if not locations:
                        continue
                    encoding = face_recognition.face_encodings(image, locations)[0]
                    person_encodings.append(encoding)
                except Exception:
                    continue

            if person_encodings:
                avg_encoding = np.mean(person_encodings, axis=0)
                self.known_face_encodings.append(avg_encoding)
                self.known_face_names.append(person_name)

        self._dedupe_known_faces()
        self._save_cache()

    def detect_and_recognize_faces(self, frame, model="hog", tolerance=0.6, device="cpu", far_mode=False):
        rgb_frame = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
        upsample_times = 1 if far_mode else 0
        try:
            face_locations = face_recognition.face_locations(
                rgb_frame,
                model=model,
                number_of_times_to_upsample=upsample_times,
            )
        except Exception:
            face_locations = face_recognition.face_locations(
                rgb_frame,
                model="hog",
                number_of_times_to_upsample=upsample_times,
            )

        # Extra pass for distant/small faces.
        if far_mode and not face_locations:
            try:
                scale_up = 1.5
                big_frame = cv2.resize(rgb_frame, None, fx=scale_up, fy=scale_up, interpolation=cv2.INTER_LINEAR)
                big_locations = face_recognition.face_locations(
                    big_frame,
                    model=model,
                    number_of_times_to_upsample=1,
                )
                if big_locations:
                    face_locations = [
                        (
                            int(top / scale_up),
                            int(right / scale_up),
                            int(bottom / scale_up),
                            int(left / scale_up),
                        )
                        for (top, right, bottom, left) in big_locations
                    ]
            except Exception:
                pass

        if not face_locations:
            return []

        face_encodings = face_recognition.face_encodings(rgb_frame, face_locations)
        results = []

        for (top, right, bottom, left), face_encoding in zip(face_locations, face_encodings):
            if self.known_face_encodings:
                distances = face_recognition.face_distance(self.known_face_encodings, face_encoding)
                best_idx = int(np.argmin(distances))
                best_distance = float(distances[best_idx])

                if best_distance < tolerance:
                    name = self.known_face_names[best_idx]
                    confidence = max(0.0, min(1.0, 1 - best_distance))
                else:
                    name = "Unknown"
                    confidence = 0.0
            else:
                name = "Unknown"
                confidence = 0.0

            results.append(
                {
                    "name": name,
                    "location": (top, right, bottom, left),
                    "confidence": confidence,
                }
            )

        return results

    def register_face(self, person_name, image):
        try:
            rgb_image = cv2.cvtColor(image, cv2.COLOR_BGR2RGB)
            locations = face_recognition.face_locations(rgb_image, model="hog")
            if not locations:
                return False

            person_dir = self.data_dir / person_name
            person_dir.mkdir(exist_ok=True)

            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            image_path = person_dir / f"{person_name}_{timestamp}.jpg"
            cv2.imwrite(str(image_path), image)

            self.load_known_faces()
            return True
        except Exception:
            return False

    def log_attendance(self, person_name, confidence=0.0):
     
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        excel_path = self.session_attendance_file

        try:
            if excel_path.exists():
                df = pd.read_excel(excel_path)
            else:
                df = pd.DataFrame(columns=["Timestamp", "Person", "Status", "Confidence"])

            if "Person" not in df.columns:
                df["Person"] = ""
            if "Confidence" not in df.columns:
                df["Confidence"] = "0.0%"
            if "Timestamp" not in df.columns:
                df["Timestamp"] = ""
            if "Status" not in df.columns:
                df["Status"] = "Present"

            person_mask = df["Person"] == person_name

            if person_mask.any():
                existing_str = str(df.loc[person_mask, "Confidence"].iloc[0]).strip()
                try:
                    existing_conf = float(existing_str.replace("%", "")) / 100.0
                except Exception:
                    existing_conf = 0.0

                if confidence > existing_conf:
                    df.loc[person_mask, "Timestamp"] = timestamp
                    df.loc[person_mask, "Confidence"] = f"{confidence:.1%}"
                    action = "updated"
                else:
                    action = "skipped"
            else:
                new_row = pd.DataFrame(
                    [
                        {
                            "Timestamp": timestamp,
                            "Person": person_name,
                            "Status": "Present",
                            "Confidence": f"{confidence:.1%}",
                        }
                    ]
                )
                df = pd.concat([df, new_row], ignore_index=True)
                action = "added"

            df.to_excel(excel_path, index=False, engine="openpyxl")
            self._format_attendance_sheet(excel_path)

       
            present_students = set(
                df.loc[df["Status"].astype(str).str.lower() == "present", "Person"].astype(str).str.strip().tolist()
            )
            self.export_attendance_json(present_students)
            return action
        except Exception:
            return "error"

    def export_attendance_json(self, present_students=None):
        try:
            if present_students is None:
                present_students = set()

            unique_names = sorted({str(name).strip() for name in self.known_face_names if str(name).strip()})
            students = []
            for name in unique_names:
                students.append({
                    "name": name,
                    "present": name in present_students,
                })

            payload = {
                "timestamp": datetime.now().isoformat(),
                "date": datetime.now().strftime("%Y-%m-%d"),
                "session_id": self.session_id,
                "session_file": self.session_attendance_file.name,
                "students": students,
            }

            with open(self.live_json_file, "w", encoding="utf-8") as f:
                json.dump(payload, f, indent=2)
        except Exception:
            pass

    def _format_attendance_sheet(self, excel_path):
        try:
            wb = load_workbook(excel_path)
            ws = wb.active

            ws.column_dimensions["A"].width = 20
            ws.column_dimensions["B"].width = 22
            ws.column_dimensions["C"].width = 12
            ws.column_dimensions["D"].width = 12

            header_fill = PatternFill(start_color="6B0F1A", end_color="6B0F1A", fill_type="solid")
            header_font = Font(bold=True, color="D4AF37")

            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")

            wb.save(excel_path)
        except Exception:
            pass


class AttendanceGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("School Attendance System - Facial Recognition")

        screen_w = self.root.winfo_screenwidth()
        screen_h = self.root.winfo_screenheight()
        win_w = max(1024, int(screen_w * 0.92))
        win_h = max(700, int(screen_h * 0.88))
        self.root.geometry(f"{win_w}x{win_h}")
        self.root.minsize(1024, 700)

        self.system = FacialRecognitionSystem()
        self.cap = None
        self.camera_running = False

        self.current_frame = None
        self.current_detections = []
        self.unique_logged_count = 0

        self.frame_count = 0
        self.detection_thread = None
        self.latest_detections = []
        self.auto_stop_after_id = None

        self.setup_ui()
        self.log_message("System ready")
        self.log_message(f"Loaded {len(self.system.known_face_names)} registered people")
        self.log_message(f"Session file: {self.system.session_attendance_file.name}")

    def setup_ui(self):
        style = ttk.Style()
        style.theme_use("clam")

        maroon = "#6B0F1A"
        gold = "#D4AF37"
        panel = "#F7E7C1"
        light = "#FFF8E7"

        self.root.configure(bg=maroon)
        style.configure("TFrame", background=light)
        style.configure("TLabelframe", background=panel, borderwidth=1, relief="solid")
        style.configure("TLabelframe.Label", background=panel, foreground=maroon, font=("Segoe UI", 10, "bold"))
        style.configure("TLabel", background=light, foreground=maroon)
        style.configure("TButton", background=gold, foreground=maroon, font=("Segoe UI", 9, "bold"))
        style.map("TButton", background=[("active", "#E8C766")])
        style.configure("TCombobox", fieldbackground="white", foreground=maroon)

        main = ttk.Frame(self.root)
        main.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        left = ttk.Frame(main)
        left.grid(row=0, column=0, sticky="nsew", padx=5)

        right_container = ttk.Frame(main)
        right_container.grid(row=0, column=1, sticky="nsew", padx=5)

        self.right_canvas = tk.Canvas(right_container, highlightthickness=0, bg=light)
        self.right_scrollbar = ttk.Scrollbar(right_container, orient="vertical", command=self.right_canvas.yview)
        self.right_canvas.configure(yscrollcommand=self.right_scrollbar.set)

        self.right_canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        self.right_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        self.right_frame = ttk.Frame(self.right_canvas)
        self.right_window = self.right_canvas.create_window((0, 0), window=self.right_frame, anchor="nw")

        self.right_frame.bind("<Configure>", self._on_right_frame_configure)
        self.right_canvas.bind("<Configure>", self._on_right_canvas_configure)
        self.right_frame.bind("<Enter>", self._bind_mousewheel)
        self.right_frame.bind("<Leave>", self._unbind_mousewheel)

        right = self.right_frame

        main.columnconfigure(0, weight=3)
        main.columnconfigure(1, weight=2)
        main.rowconfigure(0, weight=1)

        ttk.Label(left, text="Live Camera", font=("Arial", 12, "bold")).pack(anchor=tk.W)
        self.video_label = tk.Label(left, bg="black")
        self.video_label.pack(fill=tk.BOTH, expand=True, pady=10, padx=6)

        ctrl = ttk.LabelFrame(right, text="Camera Controls", padding=10)
        ctrl.pack(fill=tk.X, pady=5)

        self.available_cameras = []
        self.camera_var = tk.StringVar(value="Camera 0")
        camera_row = ttk.Frame(ctrl)
        camera_row.pack(fill=tk.X, pady=3)
        ttk.Label(camera_row, text="Camera:").pack(side=tk.LEFT)
        self.camera_combo = ttk.Combobox(
            camera_row,
            textvariable=self.camera_var,
            state="readonly",
            width=18,
        )
        self.camera_combo.pack(side=tk.LEFT, padx=6, fill=tk.X, expand=True)
        ttk.Button(camera_row, text="Refresh", command=self.refresh_camera_list).pack(side=tk.LEFT)

        ttk.Button(ctrl, text="Start Camera", command=self.start_camera).pack(fill=tk.X, pady=3)
        ttk.Button(ctrl, text="Stop Camera", command=self.stop_camera).pack(fill=tk.X, pady=3)
        ttk.Button(ctrl, text="Capture Frame", command=self.capture_frame).pack(fill=tk.X, pady=3)
        ttk.Button(ctrl, text="Refresh Face Data", command=self.refresh_faces).pack(fill=tk.X, pady=3)
        ttk.Button(ctrl, text="Open Session Excel", command=self.open_attendance_file).pack(fill=tk.X, pady=3)

        settings = ttk.LabelFrame(right, text="Settings", padding=10)
        settings.pack(fill=tk.X, pady=5)

        ttk.Label(settings, text="Detection Model:").pack(anchor=tk.W)
        self.model_var = tk.StringVar(value="hog")
        ttk.Combobox(
            settings,
            textvariable=self.model_var,
            values=["hog", "cnn"],
            state="readonly",
        ).pack(fill=tk.X, pady=3)

        ttk.Label(settings, text="Processor:").pack(anchor=tk.W)
        self.device_var = tk.StringVar(value="cpu")
        ttk.Radiobutton(settings, text="CPU", variable=self.device_var, value="cpu").pack(anchor=tk.W)
        ttk.Radiobutton(settings, text="GPU (if CUDA dlib installed)", variable=self.device_var, value="gpu").pack(anchor=tk.W)

        ttk.Label(settings, text="Sensitivity (0.1-1.0):").pack(anchor=tk.W)
        self.tolerance_var = tk.DoubleVar(value=0.4)
        ttk.Scale(settings, from_=0.1, to=1.0, variable=self.tolerance_var, orient=tk.HORIZONTAL).pack(fill=tk.X, pady=3)

        self.far_mode_var = tk.BooleanVar(value=False)
        ttk.Checkbutton(
            settings,
            text="Far Distance Mode (better for small/far faces)",
            variable=self.far_mode_var,
        ).pack(anchor=tk.W, pady=3)

        perf = ttk.LabelFrame(right, text="Performance", padding=10)
        perf.pack(fill=tk.X, pady=5)

        ttk.Label(perf, text="Frame Skip (process every Nth):").pack(anchor=tk.W)
        self.skip_var = tk.IntVar(value=2)
        ttk.Scale(perf, from_=1, to=4, variable=self.skip_var, orient=tk.HORIZONTAL).pack(fill=tk.X, pady=3)

        ttk.Label(perf, text="Detection Scale:").pack(anchor=tk.W)
        self.scale_var = tk.DoubleVar(value=0.75)
        ttk.Scale(perf, from_=0.5, to=1.0, variable=self.scale_var, orient=tk.HORIZONTAL).pack(fill=tk.X, pady=3)
        ttk.Button(perf, text="Apply Fast Classroom Preset", command=self.apply_fast_preset).pack(fill=tk.X, pady=4)

        timer = ttk.LabelFrame(right, text="Camera Timer", padding=10)
        timer.pack(fill=tk.X, pady=5)

        self.auto_stop_enabled_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(
            timer,
            text="Auto-stop camera",
            variable=self.auto_stop_enabled_var,
        ).pack(anchor=tk.W, pady=2)

        timer_row = ttk.Frame(timer)
        timer_row.pack(fill=tk.X, pady=2)
        ttk.Label(timer_row, text="Stop after (minutes):").pack(side=tk.LEFT)
        self.auto_stop_minutes_var = tk.StringVar(value="15")
        self.auto_stop_spin = tk.Spinbox(
            timer_row,
            from_=1,
            to=240,
            textvariable=self.auto_stop_minutes_var,
            width=6,
        )
        self.auto_stop_spin.pack(side=tk.LEFT, padx=8)
        ttk.Label(timer, text="Default: 15 min. Turn off to run continuously.").pack(anchor=tk.W)

        register = ttk.LabelFrame(right, text="Register New Face", padding=10)
        register.pack(fill=tk.X, pady=5)

        ttk.Label(register, text="Person Name:").pack(anchor=tk.W)
        self.name_var = tk.StringVar()
        ttk.Entry(register, textvariable=self.name_var).pack(fill=tk.X, pady=3)
        ttk.Button(register, text="Capture & Register", command=self.register_face).pack(fill=tk.X, pady=5)

        status = ttk.LabelFrame(right, text="Status Log", padding=10)
        status.pack(fill=tk.BOTH, expand=True, pady=5)

        self.status_text = scrolledtext.ScrolledText(status, height=12, width=40, font=("Courier", 9))
        self.status_text.pack(fill=tk.BOTH, expand=True)
        self.status_text.configure(bg="#2B0A12", fg=gold, insertbackground=gold)

        self.attendance_label = tk.Label(right, text="Not recording", fg=maroon, bg=light, font=("Segoe UI", 9, "bold"))
        self.attendance_label.pack(anchor=tk.W, pady=4)

        self.refresh_camera_list()

    def log_message(self, msg):
        timestamp = datetime.now().strftime("%H:%M:%S")
        self.status_text.insert(tk.END, f"[{timestamp}] {msg}\n")
        self.status_text.see(tk.END)
        self.root.update_idletasks()

    def _on_right_frame_configure(self, _event=None):
        self.right_canvas.configure(scrollregion=self.right_canvas.bbox("all"))

    def _on_right_canvas_configure(self, event):
        self.right_canvas.itemconfigure(self.right_window, width=event.width)

    def _on_mousewheel(self, event):
        if os.name == "nt":
            self.right_canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")
        else:
            delta = event.delta if hasattr(event, "delta") else 0
            if delta != 0:
                self.right_canvas.yview_scroll(int(-1 * delta), "units")

    def _on_mousewheel_linux_up(self, _event):
        self.right_canvas.yview_scroll(-1, "units")

    def _on_mousewheel_linux_down(self, _event):
        self.right_canvas.yview_scroll(1, "units")

    def _bind_mousewheel(self, _event=None):
        self.right_canvas.bind_all("<MouseWheel>", self._on_mousewheel)
        self.right_canvas.bind_all("<Button-4>", self._on_mousewheel_linux_up)
        self.right_canvas.bind_all("<Button-5>", self._on_mousewheel_linux_down)

    def _unbind_mousewheel(self, _event=None):
        self.right_canvas.unbind_all("<MouseWheel>")
        self.right_canvas.unbind_all("<Button-4>")
        self.right_canvas.unbind_all("<Button-5>")

    def apply_fast_preset(self):
      
        self.model_var.set("hog")
        self.skip_var.set(3)
        self.scale_var.set(0.6)
        self.log_message("Fast preset applied: HOG, frame skip 3, scale 0.60")

    def _list_available_cameras(self, max_index=8):
        found = []
        for idx in range(max_index + 1):
            cap = None
            try:
                if os.name == "nt":
                    cap = cv2.VideoCapture(idx, cv2.CAP_DSHOW)
                    if not cap.isOpened():
                        cap.release()
                        cap = cv2.VideoCapture(idx, cv2.CAP_MSMF)
                else:
                    cap = cv2.VideoCapture(idx)

                if cap is not None and cap.isOpened():
                    found.append(idx)
            except Exception:
                pass
            finally:
                if cap is not None:
                    cap.release()

        if not found:
            found = [0]
        return found

    def refresh_camera_list(self):
        self.available_cameras = self._list_available_cameras()
        values = [f"Camera {idx}" for idx in self.available_cameras]
        self.camera_combo["values"] = values
        if self.camera_var.get() not in values:
            self.camera_var.set(values[0])

    def _selected_camera_index(self):
        value = self.camera_var.get().strip()
        try:
            return int(value.split()[-1])
        except Exception:
            return 0

    def _cancel_auto_stop_timer(self):
        if self.auto_stop_after_id is not None:
            try:
                self.root.after_cancel(self.auto_stop_after_id)
            except Exception:
                pass
            self.auto_stop_after_id = None

    def _schedule_auto_stop_timer(self):
        self._cancel_auto_stop_timer()

        if not self.auto_stop_enabled_var.get():
            self.log_message("Auto-stop disabled")
            return

        try:
            minutes = int(self.auto_stop_minutes_var.get())
        except Exception:
            minutes = 15

        minutes = max(1, minutes)
        ms = minutes * 60 * 1000
        self.auto_stop_after_id = self.root.after(ms, self._auto_stop_camera)
        self.log_message(f"Auto-stop set to {minutes} minute(s)")

    def _auto_stop_camera(self):
        self.auto_stop_after_id = None
        if self.camera_running:
            self.log_message("Camera auto-stopped by timer")
            self.stop_camera()
            messagebox.showinfo("Camera Timer", "Camera stopped automatically after the set duration.")

    def start_camera(self):
        if self.camera_running:
            return

        self.refresh_camera_list()
        preferred = self._selected_camera_index()
        candidates = [preferred] + [idx for idx in self.available_cameras if idx != preferred]

        opened_cap = None
        opened_idx = None
        for idx in candidates:
            backends = [None]
            if os.name == "nt":
                backends = [cv2.CAP_DSHOW, cv2.CAP_MSMF, None]

            for backend in backends:
                try:
                    cap = cv2.VideoCapture(idx, backend) if backend is not None else cv2.VideoCapture(idx)
                    if cap.isOpened():
                        opened_cap = cap
                        opened_idx = idx
                        break
                    cap.release()
                except Exception:
                    continue
            if opened_cap is not None:
                break

        self.cap = opened_cap
        if self.cap is None or not self.cap.isOpened():
            messagebox.showerror("Error", "Could not open camera")
            self.cap = None
            return

        self.cap.set(cv2.CAP_PROP_BUFFERSIZE, 1)
        self.cap.set(cv2.CAP_PROP_FRAME_WIDTH, 960)
        self.cap.set(cv2.CAP_PROP_FRAME_HEIGHT, 540)

        self.camera_running = True
        self.unique_logged_count = 0
        self.attendance_label.config(text="Recording - 0 unique logged", fg="green")
        self.log_message(f"Camera started (Camera {opened_idx})")
        self._schedule_auto_stop_timer()
        self.update_frame()

    def stop_camera(self):
        self._cancel_auto_stop_timer()
        self.camera_running = False
        if self.cap:
            self.cap.release()
            self.cap = None
        self.attendance_label.config(text="Stopped", fg="gray")
        self.log_message("Camera stopped")

    def _detect_faces_async(self, frame, scale):
        try:
            effective_scale = scale
            if self.far_mode_var.get() and effective_scale < 0.9:
                effective_scale = 0.9

            if effective_scale < 1.0:
                small = cv2.resize(
                    frame,
                    (int(frame.shape[1] * effective_scale), int(frame.shape[0] * effective_scale)),
                )
            else:
                small = frame

            detections = self.system.detect_and_recognize_faces(
                small,
                model=self.model_var.get(),
                tolerance=self.tolerance_var.get(),
                device=self.device_var.get(),
                far_mode=self.far_mode_var.get(),
            )

            if effective_scale < 1.0 and detections:
                inv = 1.0 / effective_scale
                for det in detections:
                    top, right, bottom, left = det["location"]
                    det["location"] = (
                        int(top * inv),
                        int(right * inv),
                        int(bottom * inv),
                        int(left * inv),
                    )

            self.latest_detections = detections
        except Exception:
            self.latest_detections = []

    def update_frame(self):
        if not (self.camera_running and self.cap):
            return

        ret, frame = self.cap.read()
        if not ret:
            self.root.after(30, self.update_frame)
            return

        self.frame_count += 1
        frame_skip = max(1, int(self.skip_var.get()))
        scale = float(self.scale_var.get())

        if self.frame_count % frame_skip == 0:
            if not (self.detection_thread and self.detection_thread.is_alive()):
                self.detection_thread = threading.Thread(
                    target=self._detect_faces_async,
                    args=(frame.copy(), scale),
                    daemon=True,
                )
                self.detection_thread.start()

        detections = self.latest_detections
        display = frame.copy()

        for det in detections:
            top, right, bottom, left = det["location"]
            name = det["name"]
            conf = det["confidence"]

            color = (0, 255, 0) if name != "Unknown" else (0, 0, 255)
            cv2.rectangle(display, (left, top), (right, bottom), color, 2)

            label = f"{name} {conf:.0%}" if conf else name
            cv2.rectangle(display, (left, top - 35), (right, top), color, cv2.FILLED)
            cv2.putText(
                display,
                label,
                (left + 5, top - 10),
                cv2.FONT_HERSHEY_SIMPLEX,
                0.6,
                (255, 255, 255),
                2,
            )

            if name != "Unknown":
                result = self.system.log_attendance(name, conf)
                if result == "added":
                    self.unique_logged_count += 1
                    self.attendance_label.config(text=f"Recording - {self.unique_logged_count} unique logged", fg="green")
                    self.log_message(f"+ {name} logged ({conf:.0%})")
                elif result == "updated":
                    self.log_message(f"* {name} confidence updated ({conf:.0%})")

        self.current_frame = display.copy()
        self.current_detections = detections

        rgb = cv2.cvtColor(display, cv2.COLOR_BGR2RGB)

        box_w = max(320, self.video_label.winfo_width())
        box_h = max(240, self.video_label.winfo_height())
        frame_h, frame_w = rgb.shape[:2]
        ratio = min(box_w / frame_w, box_h / frame_h)
        new_w = max(1, int(frame_w * ratio))
        new_h = max(1, int(frame_h * ratio))
        resized = cv2.resize(rgb, (new_w, new_h))

        canvas = np.zeros((box_h, box_w, 3), dtype=np.uint8)
        y0 = (box_h - new_h) // 2
        x0 = (box_w - new_w) // 2
        canvas[y0 : y0 + new_h, x0 : x0 + new_w] = resized

        photo = ImageTk.PhotoImage(Image.fromarray(canvas))
        self.video_label.config(image=photo)
        self.video_label.image = photo

        self.root.after(30, self.update_frame)

    def capture_frame(self):
        if self.current_frame is None:
            messagebox.showwarning("Warning", "No frame to capture")
            return

        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        out_path = self.system.base_dir / f"captured_{ts}.jpg"
        cv2.imwrite(str(out_path), self.current_frame)
        self.log_message(f"Frame saved: {out_path.name}")

    def register_face(self):
        person_name = self.name_var.get().strip()
        if not person_name:
            messagebox.showwarning("Warning", "Please enter a person name")
            return
        if self.current_frame is None:
            messagebox.showwarning("Warning", "No frame available")
            return

        ok = self.system.register_face(person_name, self.current_frame)
        if ok:
            self.log_message(f"Face registered for {person_name}")
            self.name_var.set("")
            messagebox.showinfo("Success", f"Face registered for {person_name}")
        else:
            messagebox.showerror("Error", "Could not register face. Make sure a clear face is visible.")

    def refresh_faces(self):
        self.log_message("Refreshing face database...")

        def worker():
            self.system.load_known_faces()
            self.system.export_attendance_json(set())
            self.log_message(f"Loaded {len(self.system.known_face_names)} registered people")

        threading.Thread(target=worker, daemon=True).start()

    def open_attendance_file(self):
        excel_path = self.system.session_attendance_file

        if not excel_path.exists():
            messagebox.showwarning("No File", "No attendance file for this session yet.")
            return

        try:
            if os.name == "nt":
                os.startfile(str(excel_path))
            else:
                messagebox.showinfo("Path", str(excel_path))
        except Exception as e:
            messagebox.showerror("Error", f"Could not open file: {e}")


def main():
    root = tk.Tk()
    app = AttendanceGUI(root)

    def on_close():
        app.stop_camera()
        root.destroy()

    root.protocol("WM_DELETE_WINDOW", on_close)
    root.mainloop()


if __name__ == "__main__":
    main()
